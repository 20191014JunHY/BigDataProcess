from openpyxl import load_workbook
wb = load_workbook(filename = 'C:\git\BigDataProcess\HW2\student.xlsx')
ws = wb.active

count = 1
while True:                                 #총학생수 구하기
    if ws['G' + str(count)].value == None : break
    count += 1
count -= 2

score_list = []                             #동점덩어리 점수 구하기
score_count = 1
for i in range(2, count):
    if ws['G' + str(count)].value == ws['G' + str(count)].value:
        score_count += 1
    else:
        score_list.append(score_count)
        print(score_list)
        score_count = 1
score_list.append(score_count)


score_A = int(count * 0.3)                  #에이 받을 수 있는 동점덩어리 구하기
canA = 0
for i in range(0, len(score_list)): 
     if canA + score_list[i] > score_A:
        break
     canA += score_list[i]

canAp = 0
if canA != 0:
    score_Ap = int(canA / 2)                    #에이 받는 학생 중 에이플 받을 수 있는 동점덩어리 구하기
    for i in range(0, len(score_list)): 
        if canAp + score_list[i] > score_Ap:
            break
        canAp += score_list[i]
    canA -= canAp             

score_B = int(count * 0.7)                  #비 받을 수 있는 동점덩어리 구하기
canB = 0
for i in range(0, len(score_list)): 
     if canB + score_list[i] > score_B:
        break
     canB += score_list[i]
canB -= canA + canAp

canBp = 0
if canB != 0:
    score_Bp = int(canB / 2)                    #비 받는 학생 중 비플 받을 수있는 동점덩어리 구하기
    for i in range(0, len(score_list)): 
        if canBp + score_list[i] > score_Bp:
            break
        canBp += score_list[i]
    canB -= canBp

canC = count - (canAp + canA + canBp + canB)
score_Cp = int(canC / 2) / 2 #씨 받는 학생 중 씨플 받을 수 있는 동점덩어리 구하기
canCp = 0
for i in range(0, len(score_list)): 
     if canCp + score_list[i] > score_Cp + canAp + canA + canBp + canB:
        break
     canCp += score_list[i]
canCp -= canAp + canA + canBp + canB
#점수 부여하기

for i in range(2, count + 2): 
        if i < canAp + 2:
            ws['H' + str(i)] = 'A+'
        elif i < canAp + canA + 2 and i >= canAp + 2:
            ws['H' + str(i)] = 'A0'
        elif i < canAp + canA + canBp + 2 and i >= canAp + canA + 2:
            ws['H' + str(i)] = 'B+'
        elif i < canAp + canA + canBp + canB + 2 and i >= canAp + canA + canBp + 2:
            ws['H' + str(i)] = 'B0' 
        elif i < canAp + canA + canBp + canB + canCp + 2 and i >= canAp + canA + canBp + canB + 2:
            ws['H' + str(i)] = 'C+'
        else:
            ws['H' + str(i)] = 'C0'
wb.save(filename = 'student_grade.xlsx')

wb.close